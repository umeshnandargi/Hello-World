import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_wkbk(filename, sheetname):
    wb = xl.load_workbook(filename)
    sheet = wb[sheetname]
    # cell=sheet.cell(1,1)
    # print(cell.value)

    corrected_headcell = sheet.cell(1,4)
    corrected_headcell.value = "corrected_price"
    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save('transaction2.xlsx')


filename = input("enter file name: ")
inp = input("enter Sheet name:")
sheetname = inp.replace(inp[0], "S")
# print(sheetname)
process_wkbk(filename, sheetname)
